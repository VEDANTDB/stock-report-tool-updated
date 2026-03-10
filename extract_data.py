"""
extract_data.py
───────────────
Reads a screener.in XLSX file (mandatory) and optionally up to 2 PDFs
(annual report, earnings transcript / investor presentation).

Usage:
    python3 extract_data.py \
        --xlsx  path/to/Company.xlsx \
        --pdf1  path/to/AnnualReport.pdf \
        --pdf2  path/to/EarningsTranscript.pdf \
        --out   extracted_data.json

Outputs a JSON file consumed by generate_report.js
"""

import argparse, json, os, re, sys
import openpyxl
from datetime import datetime


# ─── helpers ──────────────────────────────────────────────────────────────────

def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime("%b %Y")
    return str(v) if v else ""


def pct(num, den):
    if num is None or den is None or den == 0:
        return None
    return round(num / den * 100, 1)


def growth(new, old):
    if new is None or old is None or old == 0:
        return None
    return round((new / old - 1) * 100, 1)


# ─── Excel parser ─────────────────────────────────────────────────────────────

def parse_excel(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ds = wb["Data Sheet"]

    def row_values(row_num):
        return [ds.cell(row=row_num, column=c).value for c in range(1, 12)]

    def col_values(row_num, start_col=5, end_col=11):
        """Get numeric values from columns (screener uses cols 5-11 for annual, 1-10 for Q)"""
        return [safe_float(ds.cell(row=row_num, column=c).value) for c in range(start_col, end_col + 1)]

    # ── META ──────────────────────────────────────────────────────────────────
    company_name = ds.cell(row=1, column=2).value or "Company"
    face_value   = safe_float(ds.cell(row=7, column=2).value)
    cmp          = safe_float(ds.cell(row=8, column=2).value)
    mkt_cap      = safe_float(ds.cell(row=9, column=2).value)
    shares_outstanding = round(mkt_cap / cmp, 2) if cmp and mkt_cap else None

    # ── ANNUAL P&L  (rows 16-31, cols 5-11 = up to 7 years) ──────────────────
    annual_dates   = [fmt_date(ds.cell(row=16, column=c).value) for c in range(5, 12)]
    # strip blank trailing years
    annual_dates_clean = [d for d in annual_dates if d]
    n_ann = len(annual_dates_clean)
    col_start = 11 - n_ann + 1  # first data column

    def ann_row(r):
        return [safe_float(ds.cell(row=r, column=c).value) for c in range(col_start, 12)]

    ann_sales    = ann_row(17)
    ann_rawmat   = ann_row(18)
    ann_invent   = ann_row(19)
    ann_power    = ann_row(20)
    ann_mfg      = ann_row(21)
    ann_emp      = ann_row(22)
    ann_sell     = ann_row(23)
    ann_otherexp = ann_row(24)
    ann_othinc   = ann_row(25)
    ann_depn     = ann_row(26)
    ann_int      = ann_row(27)
    ann_pbt      = ann_row(28)
    ann_tax      = ann_row(29)
    ann_pat      = ann_row(30)
    ann_div      = ann_row(31)

    # derive operating profit = sales - expenses (expenses = sum of cost lines)
    ann_opprofit = []
    for i in range(n_ann):
        s = ann_sales[i]
        costs = sum(filter(None, [ann_rawmat[i], ann_invent[i], ann_power[i],
                                   ann_mfg[i], ann_emp[i], ann_sell[i], ann_otherexp[i]]))
        ann_opprofit.append(round(s - costs, 2) if s else None)

    ann_opm  = [pct(ann_opprofit[i], ann_sales[i]) for i in range(n_ann)]
    ann_npm  = [pct(ann_pat[i], ann_sales[i]) for i in range(n_ann)]

    # EPS from shares (row 93 has adjusted shares in Cr)
    adj_shares_row = [safe_float(ds.cell(row=93, column=c).value) for c in range(col_start, 12)]
    ann_eps = []
    for i in range(n_ann):
        sh = adj_shares_row[i]
        pat = ann_pat[i]
        if sh and pat and sh > 0:
            ann_eps.append(round(pat / sh, 2))
        else:
            ann_eps.append(None)

    # ── QUARTERLY P&L (rows 41-50, cols 1-10 = up to 10 quarters) ────────────
    q_dates   = [fmt_date(ds.cell(row=41, column=c).value) for c in range(1, 11)]
    q_dates   = [d for d in q_dates if d]
    n_q = len(q_dates)

    def q_row(r):
        return [safe_float(ds.cell(row=r, column=c).value) for c in range(1, n_q + 1)]

    q_sales   = q_row(42)
    q_exp     = q_row(43)
    q_othinc  = q_row(44)
    q_depn    = q_row(45)
    q_int     = q_row(46)
    q_pbt     = q_row(47)
    q_tax     = q_row(48)
    q_pat     = q_row(49)
    q_opprof  = q_row(50)
    q_opm     = [pct(q_opprof[i], q_sales[i]) for i in range(n_q)]

    # ── BALANCE SHEET (rows 56-69, same col logic as annual) ──────────────────
    bs_dates_raw = [ds.cell(row=56, column=c).value for c in range(5, 12)]
    bs_dates = [fmt_date(v) for v in bs_dates_raw if v]
    n_bs = len(bs_dates)
    bs_col_start = 11 - n_bs + 1

    def bs_row(r):
        return [safe_float(ds.cell(row=r, column=c).value) for c in range(bs_col_start, 12)]

    bs_eq_cap  = bs_row(57)
    bs_res     = bs_row(58)
    bs_borr    = bs_row(59)
    bs_othliab = bs_row(60)
    bs_total   = bs_row(61)
    bs_netblk  = bs_row(62)
    bs_cwip    = bs_row(63)
    bs_invest  = bs_row(64)
    bs_othast  = bs_row(65)
    bs_recv    = bs_row(67)
    bs_inv     = bs_row(68)
    bs_cash    = bs_row(69)

    bs_equity  = [round((bs_eq_cap[i] or 0) + (bs_res[i] or 0), 2) for i in range(n_bs)]
    bs_debt_eq = [round(bs_borr[i] / bs_equity[i], 2) if bs_equity[i] else None for i in range(n_bs)]

    # ROE
    roe = []
    for i in range(n_bs):
        eq = bs_equity[i]
        pat = ann_pat[i] if i < n_ann else None
        roe.append(pct(pat, eq) if eq and pat else None)

    # Debtor days
    ddays = []
    for i in range(n_bs):
        s = ann_sales[i] if i < n_ann else None
        r = bs_recv[i]
        ddays.append(round(r / (s / 365), 1) if s and r and s > 0 else None)

    # ── CASH FLOW (rows 81-85) ────────────────────────────────────────────────
    cf_dates_raw = [ds.cell(row=81, column=c).value for c in range(5, 12)]
    cf_dates = [fmt_date(v) for v in cf_dates_raw if v]
    n_cf = len(cf_dates)
    cf_col = 11 - n_cf + 1

    def cf_row(r):
        return [safe_float(ds.cell(row=r, column=c).value) for c in range(cf_col, 12)]

    cf_ops  = cf_row(82)
    cf_inv  = cf_row(83)
    cf_fin  = cf_row(84)
    cf_net  = cf_row(85)

    # ── PRICE HISTORY (row 90) ────────────────────────────────────────────────
    price_hist = [safe_float(ds.cell(row=90, column=c).value) for c in range(5, 12)]

    # ── DERIVED RATIOS ────────────────────────────────────────────────────────
    trailing_pe = round(cmp / ann_eps[-1], 1) if cmp and ann_eps and ann_eps[-1] else None
    trailing_ps = round(mkt_cap / ann_sales[-1], 2) if mkt_cap and ann_sales and ann_sales[-1] else None

    # ── SALES CAGR ────────────────────────────────────────────────────────────
    cagr_3y = growth(ann_sales[-1], ann_sales[-4]) / 3 if len(ann_sales) >= 4 and ann_sales[-4] else None
    cagr_5y = growth(ann_sales[-1], ann_sales[-6]) / 5 if len(ann_sales) >= 6 and ann_sales[-6] else None
    pat_cagr_3y = growth(ann_pat[-1], ann_pat[-4]) / 3 if len(ann_pat) >= 4 and ann_pat[-4] else None

    # ── RECENT QUARTER ────────────────────────────────────────────────────────
    last_q_sales   = q_sales[-1]   if q_sales   else None
    last_q_pat     = q_pat[-1]     if q_pat     else None
    last_q_opm     = q_opm[-1]     if q_opm     else None
    last_q_label   = q_dates[-1]   if q_dates   else ""
    yoy_q_sales    = growth(q_sales[-1], q_sales[-5]) if len(q_sales) >= 5 else None
    yoy_q_pat      = growth(q_pat[-1], q_pat[-5])     if len(q_pat) >= 5   else None

    return {
        "meta": {
            "company_name": company_name,
            "face_value":   face_value,
            "cmp":          cmp,
            "mkt_cap_cr":   mkt_cap,
            "shares_cr":    shares_outstanding,
            "trailing_pe":  trailing_pe,
            "trailing_ps":  trailing_ps,
        },
        "annual": {
            "years":     annual_dates_clean,
            "sales":     ann_sales,
            "op_profit": ann_opprofit,
            "opm_pct":   ann_opm,
            "pat":       ann_pat,
            "npm_pct":   ann_npm,
            "eps":       ann_eps,
            "other_income": ann_othinc,
            "depreciation": ann_depn,
            "interest":     ann_int,
            "pbt":          ann_pbt,
            "sales_cagr_3y_pct": round(cagr_3y, 1) if cagr_3y else None,
            "sales_cagr_5y_pct": round(cagr_5y, 1) if cagr_5y else None,
            "pat_cagr_3y_pct":   round(pat_cagr_3y, 1) if pat_cagr_3y else None,
        },
        "quarterly": {
            "periods":   q_dates,
            "sales":     q_sales,
            "op_profit": q_opprof,
            "opm_pct":   q_opm,
            "pat":       q_pat,
            "last_q_label": last_q_label,
            "last_q_sales": last_q_sales,
            "last_q_pat":   last_q_pat,
            "last_q_opm":   last_q_opm,
            "yoy_q_sales_pct": yoy_q_sales,
            "yoy_q_pat_pct":   yoy_q_pat,
        },
        "balance_sheet": {
            "years":    bs_dates,
            "equity":   bs_equity,
            "borr":     bs_borr,
            "total_assets": bs_total,
            "cash":     bs_cash,
            "recv":     bs_recv,
            "inventory":bs_inv,
            "roe_pct":  roe,
            "debt_eq":  bs_debt_eq,
            "debtor_days": ddays,
            "net_block":   bs_netblk,
            "cwip":        bs_cwip,
        },
        "cash_flow": {
            "years":    cf_dates,
            "ops":      cf_ops,
            "investing":cf_inv,
            "financing":cf_fin,
            "net":      cf_net,
        },
        "price_history": price_hist,
    }


# ─── PDF text extractor (stdlib zipfile + no external deps) ───────────────────

def extract_pdf_text(path: str, max_chars: int = 8000) -> str:
    """
    Extract text from PDF using pdfminer if available,
    otherwise fall back to zipfile XML scan (for text-based PDFs).
    Returns at most max_chars characters.
    """
    # Try pdfminer.six (may be pre-installed)
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(path)
        return text[:max_chars] if text else ""
    except ImportError:
        pass

    # Try pypdf
    try:
        from pypdf import PdfReader
        reader = PdfReader(path)
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
            if sum(len(p) for p in pages) > max_chars:
                break
        return "\n".join(pages)[:max_chars]
    except ImportError:
        pass

    # Fallback: raw binary grep for readable text lines
    try:
        with open(path, "rb") as f:
            raw = f.read()
        text_chunks = re.findall(rb'[A-Za-z0-9 .,;\-\%\(\)\:\/\+]{30,}', raw)
        decoded = [c.decode("latin-1", errors="ignore") for c in text_chunks]
        return "\n".join(decoded)[:max_chars]
    except Exception:
        return ""


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx",  required=True, help="screener.in Excel file")
    parser.add_argument("--pdf1",  default="",    help="Annual Report PDF (optional)")
    parser.add_argument("--pdf2",  default="",    help="Earnings Transcript / Investor PPT PDF (optional)")
    parser.add_argument("--out",   default="extracted_data.json")
    args = parser.parse_args()

    print(f"[1/3] Parsing Excel: {args.xlsx}")
    data = parse_excel(args.xlsx)

    if args.pdf1 and os.path.exists(args.pdf1):
        print(f"[2/3] Extracting PDF1: {args.pdf1}")
        data["pdf1_text"] = extract_pdf_text(args.pdf1, max_chars=10000)
        data["pdf1_name"] = os.path.basename(args.pdf1)
    else:
        data["pdf1_text"] = ""
        data["pdf1_name"] = ""

    if args.pdf2 and os.path.exists(args.pdf2):
        print(f"[3/3] Extracting PDF2: {args.pdf2}")
        data["pdf2_text"] = extract_pdf_text(args.pdf2, max_chars=10000)
        data["pdf2_name"] = os.path.basename(args.pdf2)
    else:
        data["pdf2_text"] = ""
        data["pdf2_name"] = ""

    with open(args.out, "w") as f:
        json.dump(data, f, indent=2)

    print(f"\n✅  Data extracted → {args.out}")
    print(f"    Company  : {data['meta']['company_name']}")
    print(f"    CMP      : ₹{data['meta']['cmp']}")
    print(f"    Mkt Cap  : ₹{data['meta']['mkt_cap_cr']} Cr")
    print(f"    Annual years : {data['annual']['years']}")
    print(f"    Latest Q : {data['quarterly']['last_q_label']}")


if __name__ == "__main__":
    main()
